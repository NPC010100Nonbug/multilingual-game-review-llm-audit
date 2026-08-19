#!/usr/bin/env python3
"""07_build_hardneg_arm.py — fresh English hard-negative retrieval arm.

TWO-STAGE, gated by the user's signature.

  STAGE 1 (default, no flag): output ONLY candidate counts and the sampling
    distribution. Reads raw review text INTERNALLY to run the keyword regexes
    (the frame has no text), but surfaces NO text, writes NO file, runs NO LLM.

  STAGE 2 (--build, run only after the user says "开始标注"): using the SAME
    deterministic draw, emit
      - data/raw/hardneg_workbook.xlsx        blind labeling book (no
            keyword_class, no prior label; input_schema §2 given cols gray)
      - data/splits/pilot_prompt_extension_manifest.csv   per-row provenance
            (review_id, appid, lang, keyword_class, arm_role, source)
      - data/splits/pilot_prompt_extension_counts.csv     per-class candidate /
            drawn / shortfall record
      - data/splits/reserved_ids.csv          unified reserved-ID registry that
            every 后减前 carve (02_align_sample / 02a*) now unions in, so these
            80 can never fall into gold/dev/train/stress.
    Before AND after writing, assert the 80 ids do not intersect
    split_manifest / gold / stress. (train/dev do not exist yet -> their guard
    is the reserved_ids union already wired into the carve scripts, enforced
    whenever they are eventually run.)

Firewall (后减前): pool = English machine-eligible frame
  MINUS split_manifest (gold/pilot_draft/pilot_prompt)
  MINUS stress_candidate_manifest
  MINUS current diagnostic-arm 32 ids (redundant w/ split_manifest, named
        explicitly for defensive hygiene).
The 80 are NOT written into split_manifest (that would inflate the 135-row
pilot_prompt base-rate anchor); they live in the extension manifest + reserved
registry only.

Multi-class hits: assigned to ONE class by pre-registered priority
class1 > class2 > class3 > class4, then de-duplicated -> four disjoint pools.
"""
import csv, json, glob, re, random, argparse, os
from collections import defaultdict

ROOT = "/Users/npc001/Documents/multilingual-game-review-llm-audit"
SEED = "20260817-en-hardneg-v1"
PER_CLASS = 20
ADDED_DATE = "2026-08-17"
SOURCE = "fresh_en_retrieval_v1"
GAME_NAME = {730: "CS2", 1517290: "BF2042", 1245620: "Elden Ring"}

SPLITS = f"{ROOT}/data/splits"
RESERVED_IDS = f"{SPLITS}/reserved_ids.csv"
EXT_MANIFEST = f"{SPLITS}/pilot_prompt_extension_manifest.csv"
EXT_COUNTS = f"{SPLITS}/pilot_prompt_extension_counts.csv"
WORKBOOK = f"{ROOT}/data/raw/hardneg_workbook.xlsx"

# --- Four keyword classes (Codex spec). Priority = list order (class1 highest).
CLASSES = [
    ("class1_anticheat", re.compile(
        r"\banti.?cheat\b|\bVAC\b|\bcheat(?:er|ers|ing)?\b|\bhack(?:er|ers|ing|s)?\b"
        r"|\baimbot\b|\bmalware\b", re.I)),
    ("class2_tech_compat", re.compile(
        r"secure ?boot|\bBIOS\b|\bLinux\b|steam ?deck|\bTPM\b|\bcrash(?:es|ed|ing)?\b"
        r"|\bFPS\b|\blaunch(?:er|es|ed|ing)?\b", re.I)),
    ("class3_price_cosmetic", re.compile(
        r"battle ?pass|\bDLC\b|\bskins?\b|cosmetic|micro.?transaction|\bprices?\b"
        r"|\beditions?\b|pay.?to.?win|\bp2w\b|season ?pass", re.I)),
    ("class4_difficulty_balance", re.compile(
        r"\bhard\b|\bboss(?:es)?\b|difficult|\bOP\b|\bbalanc(?:e|ed|ing)?\b"
        r"|matchmak\w*|\bMMR\b|\branked\b", re.I)),
]
CLASS_NAMES = [c for c, _ in CLASSES]

# HEADER identical to 06_build_diagnostic_arm so labeling is the same procedure.
HEADER = ["case_no","review_id","lang","game","review","voted_up","votes_funny",
    "votes_up","received_for_free","steam_purchase","weighted_vote_score",
    "written_during_early_access","out_of_scope","unfair_label","subtype",
    "facet_cheating_governance","facet_sanction","facet_access_exclusion",
    "facet_competitive_balance","facet_unfair_by_design","evidence_span",
    "normalized_claim","explicitness","confidence","borderline",
    "uncertainty_reason","annotator_note","QC"]
GIVEN_COLS = 12  # A:L given (gray); M:AB to fill (white)


def load_split_and_stress():
    split, gold, stress = set(), set(), set()
    with open(f"{SPLITS}/split_manifest.csv") as f:
        for r in csv.DictReader(f):
            split.add(str(r["review_id"]))
            if r["role"] == "gold":
                gold.add(str(r["review_id"]))
    with open(f"{SPLITS}/stress_candidate_manifest.csv") as f:
        for r in csv.DictReader(f):
            stress.add(str(r["review_id"]))
    return split, gold, stress


def load_arm_ids():
    import openpyxl
    wb = openpyxl.load_workbook(f"{ROOT}/data/raw/diagnostic_arm_workbook.xlsx",
                                read_only=True)
    key = wb["Key (SEALED)"]
    arm = set()
    for row in key.iter_rows(min_row=4, values_only=True):
        if row and row[1] is not None:
            arm.add(str(row[1]))
    wb.close()
    return arm


def load_en_pool(exclude):
    pool = set()
    with open(f"{ROOT}/data/processed/machine_eligible_frame.csv") as f:
        for r in csv.DictReader(f):
            if r["lang"] == "en" and str(r["review_id"]) not in exclude:
                pool.add(str(r["review_id"]))
    return pool


def classify(pool, capture=None):
    """Stream raw en text; assign each pool id to ONE class by priority.
    If `capture` (a set of ids) is given, also return full raw records for them.
    """
    hits = defaultdict(dict)   # class -> {rid: appid}
    records = {}
    seen = 0
    for path in glob.glob(f"{ROOT}/data/raw/*_en.jsonl"):
        for line in open(path):
            o = json.loads(line)
            rid = str(o.get("recommendationid"))
            if rid not in pool:
                continue
            seen += 1
            txt = o.get("review") or ""
            for cname, rx in CLASSES:
                if rx.search(txt):
                    hits[cname][rid] = o.get("appid")
                    break
            if capture and rid in capture:
                records[rid] = o
    return hits, seen, records


def draw_ids(hits):
    """Deterministic per-class draw of PER_CLASS ids. Returns {class: [ids]}."""
    draws = {}
    for cname in CLASS_NAMES:
        ids = sorted(hits.get(cname, {}))
        rng = random.Random(f"{SEED}:{cname}")
        rng.shuffle(ids)
        draws[cname] = ids[:PER_CLASS]
    return draws


def stage1(hits, seen, pool, full_excl, only_arm):
    print("=== EXCLUSIONS (后减前) ===")
    print(f"  split_manifest + stress ids : {len(full_excl) - len(only_arm):,}")
    print(f"  diagnostic-arm ids not already excluded: {len(only_arm)}")
    print(f"  total distinct excluded      : {len(full_excl):,}")
    print(f"\n=== CANDIDATE POOL ===")
    print(f"  English machine-eligible after exclusion: {len(pool):,}")
    print(f"  pool ids found in raw en text            : {seen:,}")
    draws = draw_ids(hits)
    print(f"\n=== PER-CLASS CANDIDATE COUNTS (priority-assigned, disjoint) ===")
    print(f"  seed = {SEED}   target = {PER_CLASS}/class")
    for cname in CLASS_NAMES:
        n = len(hits.get(cname, {}))
        short = max(0, PER_CLASS - n)
        flag = "" if short == 0 else f"  ⚠ SHORTFALL {short}"
        print(f"  {cname:<28} candidates={n:>6}   draw={len(draws[cname]):>2}{flag}")
    print(f"\n=== SAMPLING DISTRIBUTION (drawn ids by class x game; counts only) ===")
    for cname in CLASS_NAMES:
        by_game = defaultdict(int)
        for rid in draws[cname]:
            by_game[hits[cname][rid]] += 1
        parts = ", ".join(f"{GAME_NAME.get(a,a)}={n}"
                          for a, n in sorted(by_game.items(), key=lambda x: -x[1]))
        print(f"  {cname:<28} {parts}")
    print(f"\n  TOTAL drawn = {sum(len(v) for v in draws.values())}  (target {PER_CLASS*len(CLASSES)})")
    print("\n(no text surfaced, no workbook, no LLM, nothing written — STAGE 1 only.)")


def assert_disjoint(all80, split, gold, stress, where):
    bad = (all80 & split) | (all80 & gold) | (all80 & stress)
    assert not bad, f"[{where}] {len(bad)} reserved ids intersect split/gold/stress: {sorted(bad)[:5]}"
    # classes disjoint + exactly 80 distinct
    assert len(all80) == PER_CLASS * len(CLASSES), \
        f"[{where}] expected {PER_CLASS*len(CLASSES)} distinct ids, got {len(all80)}"


def stage2(hits, pool):
    import openpyxl
    from openpyxl.styles import PatternFill, Font

    draws = draw_ids(hits)
    for cname in CLASS_NAMES:
        if len(draws[cname]) < PER_CLASS:
            raise SystemExit(f"SHORTFALL in {cname}: only {len(draws[cname])} < {PER_CLASS}; "
                             f"stage 2 aborted (widen keywords / relax exclusion first).")
    class_of = {rid: cname for cname in CLASS_NAMES for rid in draws[cname]}
    all80 = set(class_of)

    split, gold, stress = load_split_and_stress()
    assert_disjoint(all80, split, gold, stress, "PRE-WRITE")

    # fetch full raw records for exactly the 80 (now we may open their text)
    _, _, records = classify(pool, capture=all80)
    missing = all80 - set(records)
    assert not missing, f"raw record missing for {len(missing)} ids: {sorted(missing)[:5]}"

    # stable draw order for the manifests; shuffled order for the blind workbook
    ordered = [(cname, rid) for cname in CLASS_NAMES for rid in draws[cname]]
    wb_rows = ordered[:]
    random.Random(f"{SEED}:workbook").shuffle(wb_rows)

    # --- provenance manifest (has keyword_class; NOT split_manifest) ---
    with open(EXT_MANIFEST, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["review_id","appid","lang","keyword_class","arm_role","source"])
        for cname, rid in ordered:
            w.writerow([rid, records[rid].get("appid"), "en", cname,
                        "hard_negative", SOURCE])

    # --- per-class candidate / drawn / shortfall record ---
    with open(EXT_COUNTS, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["keyword_class","candidates","drawn","shortfall","seed"])
        for cname in CLASS_NAMES:
            n = len(hits.get(cname, {}))
            w.writerow([cname, n, len(draws[cname]), max(0, PER_CLASS - n), SEED])

    # --- unified reserved-ID registry (idempotent union) ---
    existing = {}
    if os.path.exists(RESERVED_IDS):
        with open(RESERVED_IDS, newline="", encoding="utf-8") as f:
            for r in csv.DictReader(f):
                existing[str(r["review_id"])] = r
    for cname, rid in ordered:
        existing[rid] = {"review_id": rid, "appid": records[rid].get("appid"),
                         "reason": "hardneg_arm", "source": SOURCE,
                         "added_date": ADDED_DATE}
    with open(RESERVED_IDS, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["review_id","appid","reason","source","added_date"])
        w.writeheader()
        for rid in sorted(existing, key=int):
            w.writerow(existing[rid])

    # --- blind workbook (no keyword_class, no prior label) ---
    book = openpyxl.Workbook()
    rm = book.active; rm.title = "Read Me"
    msg = [
        "英文硬负例臂 (English Hard-Negative Arm) — 盲法工作簿",
        "",
        "用途：只用于【调 prompt / 抓误报 (false positive)】。这批是主题像不公平、",
        "  但需人工判定是否真 PRESENT 的英文近失负例。属开发集，不进 gold/stress 报告。",
        "",
        "来源：从英文 machine-eligible 帧新抽（已扣除 split_manifest + stress + 诊断臂）。",
        f"  四类各 20 条，共 {len(ordered)} 条；乱序 (seed={SEED})；无旧标签（全新，从未标过）。",
        "",
        "标注规则：",
        "  1. 按【冻结的 codebook v1.0】逐行标；本表不显示关键词类别，避免先入为主。",
        "  2. 只填白色列 (M:AB)；灰色列 (A:L) 只读，是给定输入 (input_schema §2)。",
        "  3. 【全部保留】—— 意外标成 PRESENT / NA 也保留、不能删；那正是最有价值的边界样本。",
        "  4. 标完之前不要看 “Key (SEALED)” 表（那里有关键词类别）。",
    ]
    for i, line in enumerate(msg, 1):
        rm.cell(i, 1, line)
    rm.column_dimensions["A"].width = 110

    ws = book.create_sheet("Hard-Negative Arm")
    ws.cell(1,1,"English Hard-Negative Arm — 盲法 (blind); 关键词类别见 Key(SEALED)，标完再开")
    ws.cell(2,1,"codebook=v1.0；输入=review + input_schema §2 元数据；灰列 A:L 只读，白列 M:AB 待填")
    for c,h in enumerate(HEADER,1):
        ws.cell(4,c,h).font = Font(bold=True)
    gray = PatternFill("solid", fgColor="D9D9D9")
    for c in range(1, GIVEN_COLS+1):
        ws.cell(4,c).fill = gray
    for i,(cname,rid) in enumerate(wb_rows):
        o = records[rid]; r = 5+i
        vals = [i+1, rid, "en", GAME_NAME.get(o.get("appid"), o.get("appid")),
                o.get("review"), o.get("voted_up"), o.get("votes_funny"),
                o.get("votes_up"), o.get("received_for_free"), o.get("steam_purchase"),
                o.get("weighted_vote_score"), o.get("written_during_early_access")]
        for c,v in enumerate(vals,1):
            cell = ws.cell(r,c,v)
            if c <= GIVEN_COLS: cell.fill = gray
    ws.column_dimensions["E"].width = 70
    for col in ["B","D"]: ws.column_dimensions[col].width = 14
    ws.freeze_panes = "A5"

    key = book.create_sheet("Key (SEALED)")
    key.cell(1,1,"⚠ 标完之前不要看 —— unblinding key (关键词类别)")
    for c,h in enumerate(["case_no","review_id","keyword_class","source"],1):
        key.cell(3,c,h).font = Font(bold=True)
    for i,(cname,rid) in enumerate(wb_rows):
        key.cell(4+i,1,i+1); key.cell(4+i,2,rid)
        key.cell(4+i,3,cname); key.cell(4+i,4,SOURCE)
    key.sheet_state = "hidden"
    book.save(WORKBOOK)

    # --- POST-WRITE assertion: registry ⊇ 80, no dup, still disjoint ---
    with open(RESERVED_IDS, newline="", encoding="utf-8") as f:
        reg = [str(r["review_id"]) for r in csv.DictReader(f)]
    assert len(reg) == len(set(reg)), "reserved_ids.csv has duplicate review_ids"
    assert all80 <= set(reg), "reserved_ids.csv is missing some of the 80"
    assert_disjoint(all80, split, gold, stress, "POST-WRITE")

    print("=== STAGE 2 COMPLETE ===")
    print(f"  workbook  : {WORKBOOK}  ({len(wb_rows)} rows, blind, shuffled)")
    print(f"  provenance: {EXT_MANIFEST}")
    print(f"  counts    : {EXT_COUNTS}")
    print(f"  reserved  : {RESERVED_IDS}  (now {len(reg)} ids total)")
    print(f"  assertions: PRE-WRITE + POST-WRITE disjoint(split/gold/stress) PASSED")
    print("  firewall  : 02_align_sample + 02a* now union reserved_ids in 后减前")
    print("  (no LLM run; workbook holds raw text and stays local — data/raw gitignored)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--build", action="store_true",
                    help="STAGE 2: write workbook + manifests + reserved_ids")
    args = ap.parse_args()

    split, gold, stress = load_split_and_stress()
    arm = load_arm_ids()
    full_excl = split | stress | arm
    only_arm = arm - (split | stress)
    pool = load_en_pool(full_excl)
    hits, seen, _ = classify(pool)

    if args.build:
        stage2(hits, pool)
    else:
        stage1(hits, seen, pool, full_excl, only_arm)


if __name__ == "__main__":
    main()
