#!/usr/bin/env python3
"""06_build_diagnostic_arm.py — build the English diagnostic arm workbook.

Purpose (Codex plan §1–§4, my A–D):
  A. English PRESENT cases from pilot_draft, to be RE-LABELED by hand under
     frozen codebook v1.0 — because their existing labels are Claude's under the
     OLD v1-draft-rev6 and cannot be trusted as a human tuning target.
  B. Not renamed to the repo's `dev` role (which means model/hyperparam
     selection). This is a REFERENCE LIST only; split_manifest is NOT touched.
     Internal name: pilot_prompt_extension_positive_arm (+ hard-negative arm).
  C. English hard negatives: ABSENT reviews that share surface topics
     (cheat/hack/matchmaking/paywall/ban/difficulty ...) but were judged ABSENT,
     so the prompt can be checked for FALSE POSITIVES, not only false negatives.
  D. Built entirely from pilot_draft (already-open text) -> does NOT touch the
     gold/stress firewall and needs no prompt-freeze to open.

Blind protocol: the annotation sheet shows NO prior label / arm_role / anchor
flag. All of that lives in a sealed Key sheet, to be opened only AFTER labeling.
Rows are shuffled with a fixed seed so order carries no label signal.

Output: data/raw/diagnostic_arm_workbook.xlsx  (data/raw is gitignored; the file
holds raw review text and MUST stay local.)
"""
import json, glob, re, random
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = "/Users/npc001/Documents/multilingual-game-review-llm-audit"
SEED = 20260817
GAME_NAME = {730: "CS2", 1517290: "BF2042", 1245620: "Elden Ring"}

# surface lexicon: topics that LOOK unfair-adjacent (used to pick hard negatives)
LEX = re.compile(
    r"cheat|hack|aimbot|exploit|\bbug|glitch|matchmak|\bmmr\b|ranked|paywall|"
    r"pay.?to.?win|\bp2w\b|micro.?transaction|\bdlc\b|\bban\b|banned|refund|"
    r"difficult|\bunfair|rigged|scam|grind",
    re.I,
)

def load():
    labs = [json.loads(l) for l in open(f"{ROOT}/data/pilot/pilot_labels_claude.jsonl")]
    anchors = {str(json.loads(l)["review_id"])
               for l in open(f"{ROOT}/data/pilot/anchor_catalog.jsonl")}
    raw = {}
    for f in glob.glob(f"{ROOT}/data/raw/*_en.jsonl"):
        for l in open(f):
            o = json.loads(l); raw[str(o["recommendationid"])] = o
    return labs, anchors, raw

def build_rows(labs, anchors, raw):
    rows = []
    # A/B: English positives (all 25), keep ALL regardless of relabel outcome
    for r in labs:
        if r["language"] == "en" and r["unfair_label"] == "PRESENT":
            rid = str(r["review_id"])
            if rid not in raw:      # need the text
                continue
            rows.append(("positive", rid, rid in anchors, r["unfair_label"], raw[rid]))
    # C: English hard negatives — ABSENT + surface lexicon, exclude anchors
    for r in labs:
        if r["language"] == "en" and r["unfair_label"] == "ABSENT":
            rid = str(r["review_id"])
            if rid in anchors or rid not in raw:
                continue
            if LEX.search(raw[rid].get("review") or ""):
                rows.append(("hard_negative", rid, False, r["unfair_label"], raw[rid]))
    random.Random(SEED).shuffle(rows)
    return rows

HEADER = ["case_no","review_id","lang","game","review","voted_up","votes_funny",
    "votes_up","received_for_free","steam_purchase","weighted_vote_score",
    "written_during_early_access","out_of_scope","unfair_label","subtype",
    "facet_cheating_governance","facet_sanction","facet_access_exclusion",
    "facet_competitive_balance","facet_unfair_by_design","evidence_span",
    "normalized_claim","explicitness","confidence","borderline",
    "uncertainty_reason","annotator_note","QC"]
GIVEN_COLS = 12  # A:L given (gray); M:AB to fill (white)

def main():
    labs, anchors, raw = load()
    rows = build_rows(labs, anchors, raw)
    npos = sum(1 for x in rows if x[0]=="positive")
    nneg = sum(1 for x in rows if x[0]=="hard_negative")
    nanch = sum(1 for x in rows if x[2])
    wb = openpyxl.Workbook()

    # --- Read Me ---
    rm = wb.active; rm.title = "Read Me"
    msg = [
        "英文诊断臂 (English Diagnostic Arm) — 盲法重标工作簿",
        "",
        "用途：只用于【调 prompt / 抓漏报与误报】，不报告其准确率为最终结果。",
        "  · positive 臂：pilot_draft 里的英文 PRESENT，旧标签是 Claude 在 v1-draft-rev6 下打的，不可信。",
        "  · hard_negative 臂：主题相近但被判 ABSENT 的英文评论，用来查【误报】。",
        "",
        "标注规则：",
        "  1. 按【冻结的 codebook v1.0】标，遮住任何旧标签（本表不显示旧标签）。",
        "  2. 逐行填白色列 (M:AB)；灰色列 (A:L) 只读，是给定输入 (input_schema §2)。",
        "  3. 正例臂【全部保留】—— 即使你重标成 ABSENT/NA 也保留，那些翻掉的就是最好的边界负例。",
        "  4. 标完之前，不要打开 “Key (SEALED)” 表。",
        "",
        f"本册规模：positive {npos} 条（其中 {nanch} 条是 codebook anchor，算分时须剔除）；"
        f"hard_negative {nneg} 条。合计 {len(rows)} 条。乱序（seed={SEED}）。",
        "",
        "⚠ anchor 说明：anchor 你/模型都见过，可作教学/边界例，但不能进任何“表现分”。",
        "⚠ 数量说明：hard_negative 仅 " + str(nneg) + " 条来自 pilot_draft（随机英文 ABSENT 中含混淆主题的很少）。",
        "   若要做成 Codex 说的“配平对比组”，需再从 raw 定向补一小批英文近失负例，人工标注，",
        "   并把新 id 记录、后减前从帧中扣除（gold/stress/train 之前），此表未包含这批。",
    ]
    for i, line in enumerate(msg, 1):
        rm.cell(i, 1, line)
    rm.column_dimensions["A"].width = 110

    # --- Diagnostic Arm (blind) ---
    ws = wb.create_sheet("Diagnostic Arm")
    ws.cell(1,1,"English Diagnostic Arm — 盲法 (blind); 旧标签见 Key(SEALED)，标完再开")
    ws.cell(2,1,"codebook=v1.0；输入=review + input_schema §2 元数据；灰列 A:L 只读，白列 M:AB 待填")
    for c,h in enumerate(HEADER,1):
        cell = ws.cell(4,c,h); cell.font = Font(bold=True)
    gray = PatternFill("solid", fgColor="D9D9D9")
    for c in range(1, GIVEN_COLS+1):
        ws.cell(4,c).fill = gray
    for i,(role,rid,is_anch,prior,o) in enumerate(rows):
        r = 5+i
        vals = [i+1, rid, "en", GAME_NAME.get(o.get("appid"), o.get("appid")),
                o.get("review"), o.get("voted_up"), o.get("votes_funny"),
                o.get("votes_up"), o.get("received_for_free"), o.get("steam_purchase"),
                o.get("weighted_vote_score"), o.get("written_during_early_access")]
        for c,v in enumerate(vals,1):
            cell = ws.cell(r,c,v)
            if c <= GIVEN_COLS: cell.fill = gray
    ws.column_dimensions["E"].width = 70
    for col in ["B","D"]: ws.column_dimensions[col].width = 14
    ws.freeze_panes = "A5"

    # --- Key (SEALED) ---
    key = wb.create_sheet("Key (SEALED)")
    key.cell(1,1,"⚠ 标完之前不要看 —— unblinding key")
    kh = ["case_no","review_id","arm_role","is_anchor",
          "prior_label_claude_v1draftrev6","source"]
    for c,h in enumerate(kh,1): key.cell(3,c,h).font = Font(bold=True)
    for i,(role,rid,is_anch,prior,o) in enumerate(rows):
        key.cell(4+i,1,i+1); key.cell(4+i,2,rid); key.cell(4+i,3,role)
        key.cell(4+i,4,"YES" if is_anch else ""); key.cell(4+i,5,prior)
        key.cell(4+i,6,"pilot_draft")
    key.sheet_state = "hidden"

    out = f"{ROOT}/data/raw/diagnostic_arm_workbook.xlsx"
    wb.save(out)
    print(f"WROTE {out}")
    print(f"rows={len(rows)}  positive={npos} (anchors={nanch})  hard_negative={nneg}")
    print("case_no order is shuffled; Key sheet hidden+sealed.")

if __name__ == "__main__":
    main()
