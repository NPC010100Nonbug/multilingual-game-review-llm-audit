#!/usr/bin/env python3
"""07c_build_annotation_workbook_ja_zh.py — 日中合并盲标工作簿。

把三批行合成【一册】，格式对齐 data/raw/diagnostic_arm_annotation_workbook.xlsx
（Read Me / Annotation / QC Summary / Lists + 隐藏 Key (SEALED)）。

为什么必须合并（signed freeze/hardneg_keyword_classes_ja_zh_v1.md §6.6）：
  硬负例臂若单独成册，标注者一眼就知道「这册全是负例」，会系统性偏向 ABSENT，
  臂身份从批次归属泄漏。三批乱序合入同一册才是盲的。

三批构成（每语言）：
  hardneg    从 07b 的 manifest 读入（zh 112 / ja 104）；签署版 §3 词表 + 15 字门。
  diagnostic pilot_draft 里 Claude 判 PRESENT 的该语言全部行（zh 37 / ja 32）。
             旧标签是 Claude 在 v1-draft-rev6 下打的，不可信，必须按 codebook v1.0 重标。
  control    pilot_draft 里 Claude 判 ABSENT、且正文命中签署版 §3 词表（任一类）、
             且不在 anchor_catalog 的行，定种子随机抽 8 条。§6.6 登记的「各 8 条对照」。

诊断臂与对照臂来自 pilot_draft（正文早已开放），不动 gold/stress 防火墙，
也【不】写入 reserved_ids（它们已在 split_manifest 里）。

输出 data/raw/annotation_workbook_{lang}.xlsx —— data/raw 已 gitignored，含原始
正文，必须留在本机。
"""
import csv, json, glob, re, random, argparse, os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = "/Users/npc001/Documents/multilingual-game-review-llm-audit"
SIGNED = f"{ROOT}/freeze/hardneg_keyword_classes_ja_zh_v1.md"
GAME_NAME = {730: "CS2", 1517290: "BF2042", 1245620: "Elden Ring"}
N_CONTROL = 8

HEADER = ["case_no","review_id","lang","game","review","voted_up","votes_funny",
    "votes_up","received_for_free","steam_purchase","weighted_vote_score",
    "written_during_early_access","out_of_scope","unfair_label","subtype",
    "facet_cheating_governance","facet_sanction","facet_access_exclusion",
    "facet_competitive_balance","facet_unfair_by_design","evidence_span",
    "normalized_claim","explicitness","confidence","borderline",
    "uncertainty_reason","annotator_note","QC"]
GIVEN_COLS = 12

# 对照臂检索用；与 07b 的 §3 词表同源（此处只作 ABSENT 主题邻近筛选，不参与配额）
import importlib.util
_spec = importlib.util.spec_from_file_location("m07b", f"{ROOT}/scripts/07b_build_hardneg_arm_ja_zh.py")
_m = importlib.util.module_from_spec(_spec); _spec.loader.exec_module(_m)
LEX = {"zh": re.compile("|".join(p for _, p in _m.ZH_CLASSES), re.I),
       "ja": re.compile("|".join(p for _, p in _m.JA_CLASSES), re.I)}

WIDTHS = {"A":8,"B":16,"C":8,"D":22,"E":72,"F":10,"G":11,"H":12,"I":14,"J":14,
          "K":16,"L":32,"M":15,"N":14,"O":22,"P":18,"Q":14,"R":16,"S":18,"T":18,
          "U":38,"V":42,"W":14,"X":12,"Y":12,"Z":28,"AA":34,"AB":22}

VALIDATIONS = [
    ('"true,false"',                                          ["M","Y"]),
    ('"PRESENT,ABSENT,NA"',                                   ["N"]),
    ('",distributive,procedural,distributive;procedural"',    ["O"]),
    ('",yes"',                                                ["P","Q","R","S","T"]),
    ('",explicit,implicit"',                                  ["W"]),
    ('"high,medium,low"',                                     ["X"]),
    ('",attribution_unclear,irony_undecidable,price_boundary,'
     'technical_access_boundary,toxicity_attribution,language_cue,facet_boundary"', ["Z"]),
]


def load_raw(lang):
    raw = {}
    for f in sorted(glob.glob(f"{ROOT}/data/raw/*_{lang}.jsonl")):
        with open(f, encoding="utf-8") as fh:
            for line in fh:
                o = json.loads(line)
                raw[str(o["recommendationid"])] = o
    return raw


def build_rows(lang):
    if not os.path.exists(SIGNED):
        raise SystemExit(f"REFUSING: signed pre-registration missing: {SIGNED}")
    raw = load_raw(lang)
    rows = []

    man = f"{ROOT}/data/splits/hardneg_{lang}_manifest.csv"
    if not os.path.exists(man):
        raise SystemExit(f"REFUSING: {man} missing — run 07b --build first.")
    with open(man, newline="", encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            rid = str(r["review_id"])
            rows.append(("hardneg", rid, r["keyword_class"], "", raw[rid]))

    labs = [json.loads(l) for l in open(f"{ROOT}/data/pilot/pilot_labels_claude.jsonl", encoding="utf-8")]
    anchors = {str(json.loads(l)["review_id"])
               for l in open(f"{ROOT}/data/pilot/anchor_catalog.jsonl", encoding="utf-8")}
    for r in labs:
        if r["language"] == lang and r["unfair_label"] == "PRESENT":
            rid = str(r["review_id"])
            if rid in raw:
                rows.append(("diagnostic", rid, "", "claude_PRESENT(v1-draft-rev6)", raw[rid]))

    pool = []
    for r in labs:
        if r["language"] == lang and r["unfair_label"] == "ABSENT":
            rid = str(r["review_id"])
            if rid in anchors or rid not in raw:
                continue
            if LEX[lang].search(raw[rid].get("review") or ""):
                pool.append(rid)
    pool.sort()
    rng = random.Random(f"20260824-{lang}-control-v1")
    rng.shuffle(pool)
    for rid in pool[:N_CONTROL]:
        rows.append(("control", rid, "", "claude_ABSENT(v1-draft-rev6)", raw[rid]))

    seen, uniq = set(), []
    for t in rows:
        if t[1] in seen:
            raise SystemExit(f"duplicate review_id across arms: {t[1]}")
        seen.add(t[1]); uniq.append(t)
    random.Random(f"20260824-{lang}-annotation-merge-v1").shuffle(uniq)
    return uniq, len(pool)


def build(lang):
    rows, pool_n = build_rows(lang)
    n = len(rows)
    last = 4 + n
    label = {"zh": "简体中文", "ja": "日语"}[lang]
    n_hn = sum(1 for r in rows if r[0] == "hardneg")
    n_dg = sum(1 for r in rows if r[0] == "diagnostic")
    n_ct = sum(1 for r in rows if r[0] == "control")

    wb = openpyxl.Workbook()

    rm = wb.active; rm.title = "Read Me"
    rm.cell(1, 1, f"{label}合并臂：盲法人工标注工作簿")
    rm.cell(2, 1, "用途：只用于 prompt 调试——定位不公评论的漏报与主题近似评论的误报；"
                  "不作为最终准确率或该语言能力的证据。")
    pairs = [
        ("开始前", "1. 打开冻结的 codebook v1.0。2. 本工作簿不含旧标签与 Key；只按当前文本和 A:L 元数据判断。"),
        ("每条的顺序", "先填 out_of_scope；true 时 unfair_label 必须为 NA。否则在 PRESENT / ABSENT 之间选择。"
                    "PRESENT 必须填 subtype、evidence_span、normalized_claim、explicitness；procedural 还需至少一个 facet。"),
        ("样本构成", f"共 {n} 条{label}评论，已乱序混合，**不要试图猜某行属于哪一批**。"
                  f"批次构成只写在这里、不写进表：hard negative {n_hn} 条、诊断候选 {n_dg} 条、对照 {n_ct} 条。"
                  "三批混编是刻意的：若分册，你会知道整册都是负例而系统性偏向 ABSENT。"),
        ("全部保留", "任何一行都不许删。硬负例里被你标成 PRESENT 的、诊断候选里被你标成 ABSENT/NA 的，"
                  "正是最有价值的边界样本。"),
        ("下拉字段", "M out_of_scope；N unfair_label；O subtype；P:T 为五个 procedural facet（适用时填 yes）；"
                  "W explicitness；X confidence；Y borderline；Z uncertainty_reason。可用筛选器找未完成项。"),
        ("borderline", "borderline=true 当且仅当 confidence=low，且必须填写 uncertainty_reason。"
                     "若有多个原因，在 Z 栏用分号连接；没有不确定性时 Y=false、Z 留空。"),
        ("长评论", "正文很长的行已设固定行高，屏幕上会截断显示；点该单元格在编辑栏看全文，或临时拉高行高。正文本身完整。"),
        ("不修改的内容", "灰色列 A:L 是只读输入。不要改 review、语言、游戏或 metadata；不要新增／删除行。"),
        ("完成条件", f"每行 AB 的 QC 为 READY；QC Summary 显示 {n} 条均完成。"),
        ("后续", "完成后导出独立人工 JSONL，用于 prompt 的开发性诊断。"
               "冻结后的正式结论只来自未见的 gold；stress 单独报告。"),
        ("预登记", "硬负例批依据 freeze/hardneg_keyword_classes_ja_zh_v1.md（2026-08-24 签署）："
               "§5.1 方案 A、§5.2 长度门 ≥15 字符。"
               + ("日语 class1_anticheat 已预登记为不可得（帧内命中 96.3% 被压力集检索消耗）。" if lang == "ja" else "")),
    ]
    for i, (k, v) in enumerate(pairs, start=4):
        rm.cell(i, 1, k); rm.cell(i, 2, v).alignment = Alignment(wrap_text=True, vertical="top")
    rm.column_dimensions["A"].width = 16
    rm.column_dimensions["B"].width = 110
    for i in range(4, 4 + len(pairs)):
        rm.row_dimensions[i].height = 46

    ws = wb.create_sheet("Annotation")
    ws.cell(1, 1, f"{label} Merged Arm — {n} 条盲法人工标注（乱序；批次身份见 Key(SEALED)，标完再开）")
    ws.cell(2, 1, "codebook=v1.0；输入=review + input_schema §2 元数据；灰列 A:L 只读，白列 M:AB 待填")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    gray = PatternFill("solid", fgColor="D9D9D9")
    for c, h in enumerate(HEADER, 1):
        cell = ws.cell(4, c, h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="bottom")
        if c <= GIVEN_COLS:
            cell.fill = gray
    ws.row_dimensions[4].height = 36

    for i, (arm, rid, kcls, prior, o) in enumerate(rows):
        r = 5 + i
        txt = o.get("review") or ""
        vals = [i + 1, rid, lang, GAME_NAME.get(o.get("appid"), o.get("appid")), txt,
                o.get("voted_up"), o.get("votes_funny"), o.get("votes_up"),
                o.get("received_for_free"), o.get("steam_purchase"),
                o.get("weighted_vote_score"), o.get("written_during_early_access")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.fill = gray
            if c == 5:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(GIVEN_COLS + 1, len(HEADER) + 1):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, 28, f'=IF(COUNTA(M{r}:N{r},X{r}:Y{r})<4,"INCOMPLETE",'
                       f'IF(M{r}=TRUE,IF(N{r}="NA","READY","ERROR: NA required"),'
                       f'IF(N{r}="NA","ERROR: OOS conflict","READY")))')
        if len(txt) > 400:
            ws.row_dimensions[r].height = 220
        elif len(txt) > 120:
            ws.row_dimensions[r].height = 90

    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "F5"
    ws.auto_filter.ref = f"A4:AB{last}"

    for formula, cols in VALIDATIONS:
        dv = DataValidation(type="list", formula1=formula, allow_blank=False, showDropDown=False)
        ws.add_data_validation(dv)
        for col in cols:
            dv.add(f"{col}5:{col}{last}")

    rng = f"AB5:AB{last}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'NOT(ISERROR(SEARCH("READY",AB5)))'],
        fill=PatternFill("solid", fgColor="C6EFCE"), stopIfTrue=True))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'NOT(ISERROR(SEARCH("ERROR",AB5)))'],
        fill=PatternFill("solid", fgColor="FFC7CE"), stopIfTrue=True))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'NOT(ISERROR(SEARCH("INCOMPLETE",AB5)))'],
        fill=PatternFill("solid", fgColor="FFEB9C"), stopIfTrue=True))

    qc = wb.create_sheet("QC Summary")
    qc.cell(1, 1, f"填写进度与一致性检查：{label}合并臂").font = Font(bold=True)
    qc.cell(2, 1, "本页仅检查字段是否完整／一致，不计算模型表现，也不把该臂当作正式评估。")
    A, N_, Y, AB = f"Annotation!$A$5:$A${last}", f"Annotation!$N$5:$N${last}", \
                   f"Annotation!$Y$5:$Y${last}", f"Annotation!$AB$5:$AB${last}"
    for i, (k, f) in enumerate([
            ("总条数", f"=COUNTA({A})"), ("READY", f'=COUNTIF({AB},"READY")'),
            ("仍需检查", "=B4-B5"), ("PRESENT", f'=COUNTIF({N_},"PRESENT")'),
            ("ABSENT", f'=COUNTIF({N_},"ABSENT")'), ("NA", f'=COUNTIF({N_},"NA")'),
            ("borderline", f"=COUNTIF({Y},TRUE)")], start=4):
        qc.cell(i, 1, k); qc.cell(i, 2, f)
    qc.cell(12, 1, "按游戏"); 
    for c, h in enumerate(["游戏", "总条数", "PRESENT", "ABSENT", "NA", "READY"], 1):
        qc.cell(13, c, h).font = Font(bold=True)
    D = f"Annotation!$D$5:$D${last}"
    for i, g in enumerate(["CS2", "BF2042", "Elden Ring"], start=14):
        qc.cell(i, 1, g)
        qc.cell(i, 2, f"=COUNTIF({D},A{i})")
        for c, lab in [(3, "PRESENT"), (4, "ABSENT"), (5, "NA")]:
            qc.cell(i, c, f'=COUNTIFS({D},A{i},{N_},"{lab}")')
        qc.cell(i, 6, f'=COUNTIFS({D},A{i},{AB},"READY")')
    qc.column_dimensions["A"].width = 16
    qc.column_dimensions["B"].width = 14

    ls = wb.create_sheet("Lists")
    ls.cell(1, 1, "field").font = Font(bold=True)
    ls.cell(1, 2, "allowed values / instruction").font = Font(bold=True)
    for i, (k, v) in enumerate([
        ("out_of_scope", "true | false (true ⇒ unfair_label=NA)"),
        ("unfair_label", "PRESENT | ABSENT | NA"),
        ("subtype", "blank | distributive | procedural | distributive;procedural"),
        ("facet columns", "blank or yes; only use when subtype includes procedural"),
        ("explicitness", "blank | explicit | implicit; PRESENT only"),
        ("confidence", "high | medium | low"),
        ("borderline", "true ⇔ confidence=low"),
        ("uncertainty_reason", "blank or one code; use semicolon for multiple"),
        ("reason codes", "attribution_unclear; irony_undecidable; price_boundary; "
                         "technical_access_boundary; toxicity_attribution; language_cue; facet_boundary"),
    ], start=2):
        ls.cell(i, 1, k); ls.cell(i, 2, v)
    ls.column_dimensions["A"].width = 22
    ls.column_dimensions["B"].width = 100

    key = wb.create_sheet("Key (SEALED)")
    key.cell(1, 1, "⚠ 标完之前不要看 —— unblinding key（批次身份 / 关键词类 / 旧 Claude 标签）")
    for c, h in enumerate(["case_no", "review_id", "arm", "keyword_class", "prior_label_untrusted"], 1):
        key.cell(3, c, h).font = Font(bold=True)
    for i, (arm, rid, kcls, prior, _) in enumerate(rows):
        key.cell(4 + i, 1, i + 1); key.cell(4 + i, 2, rid)
        key.cell(4 + i, 3, arm); key.cell(4 + i, 4, kcls); key.cell(4 + i, 5, prior)
    key.column_dimensions["B"].width = 16
    key.column_dimensions["C"].width = 14
    key.column_dimensions["D"].width = 26
    key.column_dimensions["E"].width = 32
    key.sheet_state = "hidden"

    out = f"{ROOT}/data/raw/annotation_workbook_{lang}.xlsx"
    wb.save(out)
    print(f"=== {lang} ===")
    print(f"  {out}")
    print(f"  {n} 行 = hardneg {n_hn} + diagnostic {n_dg} + control {n_ct}"
          f"   (对照候选池 {pool_n} 条中抽 {n_ct})")
    return out, n, n_hn, n_dg, n_ct


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--lang", choices=["ja", "zh", "both"], default="both")
    a = ap.parse_args()
    for lang in (["zh", "ja"] if a.lang == "both" else [a.lang]):
        build(lang)


if __name__ == "__main__":
    main()
