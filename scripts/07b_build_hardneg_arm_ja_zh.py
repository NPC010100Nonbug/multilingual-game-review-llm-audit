#!/usr/bin/env python3
"""07b_build_hardneg_arm_ja_zh.py — Japanese / Simplified-Chinese hard-negative arm.

Sibling of 07_build_hardneg_arm.py (English). Same two-stage structure, same
后减前 firewall, same PRE/POST-WRITE disjointness assertions. What differs is
registered in freeze/hardneg_keyword_classes_ja_zh_v1_draft.md, not here:

  * the two keyword lexicons (§3 of that file, copied VERBATIM below);
  * §5.1 ruling = OPTION A -> the Japanese arm runs THREE classes.
    ja class1_anticheat is pre-registered as UNAVAILABLE: 96.3% of its 320 frame
    hits were consumed by stress retrieval, leaving 4. That is a corpus-size
    fact (ja frame = 1/30 of en), not a lexicon defect -- en/ja/zh class1 fire
    on 5.3% / 3.0% / 3.4% of their frames. The cross-language FP table carries
    an explicit empty cell with that reason; it is not silently dropped.
  * §5.2 ruling = LENGTH GATE 15. A candidate needs >= 15 characters of review
    text (len(text.strip()), Unicode code points, no normalization) to enter the
    pool. Without it the zh arm fills with 2-3 character reviews: the zh frame's
    median length is 8 chars, and only ~28% of zh reviews clear 20 chars. Such
    rows are not hard negatives (no topical adjacency) and inflate NA.
    The gate narrows the length confound against the English arm's actually-drawn
    median of 135 chars; it does not remove it, hence the registered post-hoc
    length-stratified sensitivity check.

TWO-STAGE, gated by the user's signature.

  STAGE 1 (default, no flag): candidate counts and the draw distribution ONLY.
    Reads raw review text INTERNALLY to run the regexes (the frame has no text),
    but surfaces NO text, writes NO file, runs NO LLM.

  STAGE 2 (--build, run only after the user says "开始标注"): same deterministic
    draw, emit the blind workbook + provenance manifest + counts + reserved_ids.
    STAGE 2 additionally refuses to run until the SIGNED pre-registration
    freeze/hardneg_keyword_classes_ja_zh_v1.md (no _draft suffix) exists --
    freezing is the maintainer's personal signature action, never the script's.

Firewall (后减前): pool = {lang} machine-eligible frame
  MINUS split_manifest (gold / pilot_draft / pilot_prompt)
  MINUS stress_candidate_manifest
  MINUS reserved_ids.csv (the 80 English hardneg rows, and anything added later)
  MINUS diagnostic-arm workbook ids (redundant with split_manifest; named
        explicitly for defensive hygiene).
Drawn rows are NOT written into split_manifest -- that would pollute the 135-row
pilot_prompt random-arm base-rate anchor.

Multi-class hits: assigned to ONE class by pre-registered priority
class1 > class2 > class3 > class4, then de-duplicated -> disjoint pools.
"""
import csv, json, glob, re, random, argparse, os
from collections import defaultdict

ROOT = "/Users/npc001/Documents/multilingual-game-review-llm-audit"
SPLITS = f"{ROOT}/data/splits"
RESERVED_IDS = f"{SPLITS}/reserved_ids.csv"
FRAME = f"{ROOT}/data/processed/machine_eligible_frame.csv"
SIGNED_PREREG = f"{ROOT}/freeze/hardneg_keyword_classes_ja_zh_v1.md"
DRAFT_PREREG = f"{ROOT}/freeze/hardneg_keyword_classes_ja_zh_v1_draft.md"

ADDED_DATE = "2026-08-24"
MIN_LEN = 15                      # §5.2 ruling
GAME_NAME = {730: "CS2", 1517290: "BF2042", 1245620: "Elden Ring"}

# --- §3 lexicons, verbatim from the pre-registration. Priority = list order. ---
ZH_CLASSES = [
    ("class1_anticheat",
     r"外挂|挂逼|挂壁|开挂|作弊|透视|锁头|自瞄|反作弊|\bDMA\b|辅助器|挂哥"),
    ("class2_tech_compat",
     r"崩溃|闪退|卡顿|掉帧|帧数|\bFPS\b|安全启动|secure ?boot|\bTPM\b|\bBIOS\b|主板|显卡驱动|黑屏|报错|启动不了|进不去|Linux|steam ?deck|优化"),
    ("class3_price_cosmetic",
     r"氪金|付费|充值|内购|微交易|\bDLC\b|皮肤|外观|通行证|战令|价格|售价|打折|\bP2W\b|pay.?to.?win|豪华版|典藏版|终极版"),
    ("class4_difficulty_balance",
     r"难度|太难|很难|\bboss\b|平衡|数值|匹配|段位|排位|强度|新手|大佬|\bOP\b"),
]
JA_CLASSES = [
    ("class1_anticheat",
     r"チート|チータ|ハッカー|ハック|エイムボット|アンチチート|不正行為|\bVAC\b|ウォールハック|野良チー"),
    ("class2_tech_compat",
     r"クラッシュ|フリーズ|強制終了|落ちる|落ちます|カクつ|カクカク|フレームレート|\bFPS\b|セキュアブート|secure ?boot|\bTPM\b|\bBIOS\b|マザーボード|Linux|steam ?deck|proton|起動しない|起動できない|エラー"),
    ("class3_price_cosmetic",
     r"課金|有料|\bDLC\b|スキン|バトルパス|シーズンパス|価格|値段|マイクロトランザクション|\bP2W\b|pay.?to.?win|デラックス|エディション|定価"),
    ("class4_difficulty_balance",
     r"難易度|難しい|難しく|むずかし|ボス|バランス|マッチング|マッチメイキング|ランクマ|ランク帯|\bOP\b|強すぎ|初心者|レート"),
]

# §5.1 OPTION A: ja class1 is retrieved and reported, but NOT drawn from.
# Keeping it in RETRIEVE lets STAGE 1 keep re-measuring the gap on every run, so
# the empty cell stays an evidenced claim rather than a remembered one.
LANG = {
    "zh": {
        "classes": ZH_CLASSES,
        "quota": {"class1_anticheat": 28, "class2_tech_compat": 28,
                  "class3_price_cosmetic": 28, "class4_difficulty_balance": 28},
        "unavailable": {},
        "seed": "20260824-zh-hardneg-v1",
        "reason": "hardneg_arm_zh",
        "source": "fresh_zh_retrieval_v1",
        "glob": "*_zh.jsonl",
    },
    "ja": {
        "classes": JA_CLASSES,
        "quota": {"class2_tech_compat": 35, "class3_price_cosmetic": 34,
                  "class4_difficulty_balance": 35},
        "unavailable": {"class1_anticheat":
                        "pre-registered UNAVAILABLE (§5.1 option A): 96.3% of 320 "
                        "frame hits consumed by stress retrieval; corpus size, not lexicon"},
        "seed": "20260824-ja-hardneg-v1",
        "reason": "hardneg_arm_ja",
        "source": "fresh_ja_retrieval_v1",
        "glob": "*_ja.jsonl",
    },
}

HEADER = ["case_no","review_id","lang","game","review","voted_up","votes_funny",
    "votes_up","received_for_free","steam_purchase","weighted_vote_score",
    "written_during_early_access","out_of_scope","unfair_label","subtype",
    "facet_cheating_governance","facet_sanction","facet_access_exclusion",
    "facet_competitive_balance","facet_unfair_by_design","evidence_span",
    "normalized_claim","explicitness","confidence","borderline",
    "uncertainty_reason","annotator_note","QC"]
GIVEN_COLS = 12  # A:L given (gray); M:AB to fill (white)


def compiled(classes):
    return [(c, re.compile(p, re.I)) for c, p in classes]


def load_exclusions():
    """split_manifest + stress_candidate_manifest + reserved_ids + diagnostic arm."""
    split, gold, stress, reserved = set(), set(), set(), set()
    with open(f"{SPLITS}/split_manifest.csv") as f:
        for r in csv.DictReader(f):
            split.add(str(r["review_id"]))
            if r["role"] == "gold":
                gold.add(str(r["review_id"]))
    with open(f"{SPLITS}/stress_candidate_manifest.csv") as f:
        for r in csv.DictReader(f):
            stress.add(str(r["review_id"]))
    if os.path.exists(RESERVED_IDS):
        with open(RESERVED_IDS) as f:
            for r in csv.DictReader(f):
                reserved.add(str(r["review_id"]))
    return split, gold, stress, reserved


def load_arm_ids():
    """Diagnostic-arm ids from the sealed key sheet (defensive; ⊆ split_manifest)."""
    path = f"{ROOT}/data/raw/diagnostic_arm_workbook.xlsx"
    if not os.path.exists(path):
        return set()
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    arm = set()
    if "Key (SEALED)" in wb.sheetnames:
        for row in wb["Key (SEALED)"].iter_rows(min_row=4, values_only=True):
            if row and len(row) > 1 and row[1] is not None:
                arm.add(str(row[1]))
    wb.close()
    return arm


def load_pool(lang, exclude):
    pool, frame_n = set(), 0
    with open(FRAME) as f:
        for r in csv.DictReader(f):
            if r["lang"] != lang:
                continue
            frame_n += 1
            rid = str(r["review_id"])
            if rid not in exclude:
                pool.add(rid)
    return pool, frame_n


def classify(lang, pool, capture=None):
    """Stream raw text; assign each pool id to ONE class by priority.

    Returns hits[class][rid] = (appid, length) for rows clearing MIN_LEN, plus
    ungated[class] counts so the gate's effect stays visible on every run.
    """
    rx = compiled(LANG[lang]["classes"])
    hits = defaultdict(dict)
    ungated = defaultdict(int)
    records, seen = {}, 0
    for path in sorted(glob.glob(f"{ROOT}/data/raw/{LANG[lang]['glob']}")):
        with open(path) as fh:
            for line in fh:
                o = json.loads(line)
                rid = str(o.get("recommendationid"))
                if rid not in pool:
                    continue
                seen += 1
                txt = o.get("review") or ""
                n = len(txt.strip())
                for cname, r in rx:
                    if r.search(txt):
                        ungated[cname] += 1
                        if n >= MIN_LEN:
                            hits[cname][rid] = (o.get("appid"), n)
                            if capture and rid in capture:
                                records[rid] = o
                        break
    return hits, ungated, seen, records


def draw_ids(lang, hits):
    """Deterministic per-class draw. Classes with no quota are never drawn from."""
    cfg = LANG[lang]
    draws = {}
    for cname in cfg["quota"]:
        ids = sorted(hits.get(cname, {}))
        rng = random.Random(f"{cfg['seed']}:{cname}")
        rng.shuffle(ids)
        draws[cname] = ids[:cfg["quota"][cname]]
    return draws


def med(vals):
    v = sorted(vals)
    return v[len(v) // 2] if v else 0


def stage1(lang, hits, ungated, seen, pool, frame_n, excl_detail):
    cfg = LANG[lang]
    total_q = sum(cfg["quota"].values())
    print(f"\n{'='*74}\n=== {lang.upper()}  hard-negative arm — STAGE 1 (counts only) ===\n{'='*74}")
    print(f"  seed = {cfg['seed']}   length gate = >= {MIN_LEN} chars   quota total = {total_q}")
    print("\n--- 后减前 EXCLUSIONS ---")
    for k, v in excl_detail:
        print(f"  {k:<44}{v:>9,}")
    print("\n--- CANDIDATE POOL ---")
    print(f"  {lang} machine-eligible frame                     {frame_n:>9,}")
    print(f"  after exclusion                              {len(pool):>9,}")
    print(f"  pool ids located in raw text                 {seen:>9,}")
    matched = sum(ungated.values())
    print(f"  matched any class (before length gate)       {matched:>9,}"
          f"   ({matched/max(seen,1):.1%} of pool)")

    print(f"\n--- PER-CLASS CANDIDATES (priority-assigned, disjoint) ---")
    print(f"  {'class':<28}{'pre-gate':>9}{'>=15':>8}{'quota':>7}{'drawn':>7}   {'median len':>10}")
    draws = draw_ids(lang, hits)
    for cname, _ in cfg["classes"]:
        pre = ungated.get(cname, 0)
        post = len(hits.get(cname, {}))
        if cname in cfg["unavailable"]:
            print(f"  {cname:<28}{pre:>9,}{post:>8,}{'—':>7}{'—':>7}   "
                  f"{med([n for _, n in hits.get(cname, {}).values()]):>10,}")
            print(f"      ⚠ {cfg['unavailable'][cname]}")
            continue
        q = cfg["quota"][cname]
        d = len(draws[cname])
        flag = "" if d == q else f"   ⚠ SHORTFALL {q - d}"
        print(f"  {cname:<28}{pre:>9,}{post:>8,}{q:>7}{d:>7}   "
              f"{med([n for _, n in hits[cname].values()]):>10,}{flag}")

    print(f"\n--- DRAW DISTRIBUTION BY GAME (counts only) ---")
    for cname in cfg["quota"]:
        by_game = defaultdict(int)
        for rid in draws[cname]:
            by_game[hits[cname][rid][0]] += 1
        parts = ", ".join(f"{GAME_NAME.get(a, a)}={n}"
                          for a, n in sorted(by_game.items(), key=lambda x: -x[1]))
        print(f"  {cname:<28} {parts}")
    drawn_len = [hits[c][r][1] for c in draws for r in draws[c]]
    print(f"\n  TOTAL drawn = {sum(len(v) for v in draws.values())}  (quota {total_q})")
    print(f"  drawn length: median={med(drawn_len):,}  min={min(drawn_len) if drawn_len else 0:,}"
          f"  max={max(drawn_len) if drawn_len else 0:,}")


def assert_disjoint(ids, split, gold, stress, reserved_before, expected_n, where):
    bad = (ids & split) | (ids & gold) | (ids & stress) | (ids & reserved_before)
    assert not bad, f"[{where}] {len(bad)} drawn ids intersect split/gold/stress/reserved: {sorted(bad)[:5]}"
    assert len(ids) == expected_n, f"[{where}] expected {expected_n} distinct ids, got {len(ids)}"


def stage2(lang, hits, pool):
    if not os.path.exists(SIGNED_PREREG):
        raise SystemExit(
            "REFUSING to build.\n"
            f"  The signed pre-registration does not exist: {SIGNED_PREREG}\n"
            f"  Only the unsigned draft is present:        {DRAFT_PREREG}\n"
            "  Freezing is the maintainer's personal signature action. Sign the\n"
            "  pre-registration first (drop the _draft suffix, fill the date), then\n"
            "  re-run with --build.")

    import openpyxl
    from openpyxl.styles import PatternFill, Font

    cfg = LANG[lang]
    draws = draw_ids(lang, hits)
    for cname, q in cfg["quota"].items():
        if len(draws[cname]) < q:
            raise SystemExit(f"SHORTFALL in {lang}/{cname}: {len(draws[cname])} < {q}; "
                             f"stage 2 aborted (re-register the lexicon or the quota first).")
    class_of = {rid: c for c in draws for rid in draws[c]}
    ids = set(class_of)
    expected_n = sum(cfg["quota"].values())

    split, gold, stress, reserved_before = load_exclusions()
    assert_disjoint(ids, split, gold, stress, reserved_before, expected_n, "PRE-WRITE")

    _, _, _, records = classify(lang, pool, capture=ids)
    missing = ids - set(records)
    assert not missing, f"raw record missing for {len(missing)} ids: {sorted(missing)[:5]}"

    ordered = [(c, rid) for c in cfg["quota"] for rid in draws[c]]
    wb_rows = ordered[:]
    random.Random(f"{cfg['seed']}:workbook").shuffle(wb_rows)

    ext_manifest = f"{SPLITS}/hardneg_{lang}_manifest.csv"
    ext_counts = f"{SPLITS}/hardneg_{lang}_counts.csv"
    workbook = f"{ROOT}/data/raw/hardneg_workbook_{lang}.xlsx"

    with open(ext_manifest, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["review_id","appid","lang","keyword_class","arm_role","source","text_len"])
        for cname, rid in ordered:
            w.writerow([rid, records[rid].get("appid"), lang, cname,
                        "hard_negative", cfg["source"], hits[cname][rid][1]])

    with open(ext_counts, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["keyword_class","candidates_ge_min_len","quota","drawn","shortfall","min_len","seed","note"])
        for cname, _ in cfg["classes"]:
            n = len(hits.get(cname, {}))
            if cname in cfg["unavailable"]:
                w.writerow([cname, n, 0, 0, 0, MIN_LEN, cfg["seed"], cfg["unavailable"][cname]])
            else:
                q = cfg["quota"][cname]
                w.writerow([cname, n, q, len(draws[cname]), max(0, q - len(draws[cname])),
                            MIN_LEN, cfg["seed"], ""])

    existing = {}
    if os.path.exists(RESERVED_IDS):
        with open(RESERVED_IDS, newline="", encoding="utf-8") as f:
            for r in csv.DictReader(f):
                existing[str(r["review_id"])] = r
    for cname, rid in ordered:
        existing[rid] = {"review_id": rid, "appid": records[rid].get("appid"),
                         "reason": cfg["reason"], "source": cfg["source"],
                         "added_date": ADDED_DATE}
    with open(RESERVED_IDS, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["review_id","appid","reason","source","added_date"])
        w.writeheader()
        for rid in sorted(existing, key=int):
            w.writerow(existing[rid])

    book = openpyxl.Workbook()
    rm = book.active; rm.title = "Read Me"
    label = {"zh": "简体中文", "ja": "日语"}[lang]
    msg = [
        f"{label}硬负例臂 — 盲法工作簿",
        "",
        "用途：只用于【调 prompt / 抓误报 (false positive)】。这批是主题像不公平、",
        "  但需人工判定是否真 PRESENT 的近失负例。属开发集，不进 gold/stress 报告。",
        "",
        f"来源：从 {lang} machine-eligible 帧新抽（已扣除 split_manifest + stress + reserved + 诊断臂）。",
        f"  共 {len(ordered)} 条；正文 ≥ {MIN_LEN} 字符；乱序 (seed={cfg['seed']})；无旧标签（全新，从未标过）。",
        "",
        "标注规则：",
        "  1. 按【冻结的 codebook v1.0】逐行标；本表不显示关键词类别，避免先入为主。",
        "  2. 只填白色列 (M:AB)；灰色列 (A:L) 只读，是给定输入 (input_schema §2)。",
        "  3. 【全部保留】—— 意外标成 PRESENT / NA 也保留、不能删；那正是最有价值的边界样本。",
        "  4. 标完之前不要看 “Key (SEALED)” 表（那里有关键词类别）。",
    ]
    if cfg["unavailable"]:
        msg += ["", "预登记说明：" + "；".join(f"{k} 不可得" for k in cfg["unavailable"])]
    for i, line in enumerate(msg, 1):
        rm.cell(i, 1, line)
    rm.column_dimensions["A"].width = 110

    ws = book.create_sheet("Hard-Negative Arm")
    ws.cell(1,1,f"{label} Hard-Negative Arm — 盲法 (blind); 关键词类别见 Key(SEALED)，标完再开")
    ws.cell(2,1,"codebook=v1.0；输入=review + input_schema §2 元数据；灰列 A:L 只读，白列 M:AB 待填")
    for c,h in enumerate(HEADER,1):
        ws.cell(4,c,h).font = Font(bold=True)
    gray = PatternFill("solid", fgColor="D9D9D9")
    for c in range(1, GIVEN_COLS+1):
        ws.cell(4,c).fill = gray
    for i,(cname,rid) in enumerate(wb_rows):
        o = records[rid]; r = 5+i
        vals = [i+1, rid, lang, GAME_NAME.get(o.get("appid"), o.get("appid")),
                o.get("review"), o.get("voted_up"), o.get("votes_funny"),
                o.get("votes_up"), o.get("received_for_free"), o.get("steam_purchase"),
                o.get("weighted_vote_score"), o.get("written_during_early_access")]
        for c,v in enumerate(vals,1):
            cell = ws.cell(r,c,v)
            if c <= GIVEN_COLS: cell.fill = gray
    ws.column_dimensions["E"].width = 70
    for col in ["B","D"]: ws.column_dimensions[col].width = 14
    ws.freeze_panes = "A5"

    key = book.create_sheet("Key (SEALED)")
    key.cell(1,1,"⚠ 标完之前不要看 —— unblinding key (关键词类别)")
    for c,h in enumerate(["case_no","review_id","keyword_class","source"],1):
        key.cell(3,c,h).font = Font(bold=True)
    for i,(cname,rid) in enumerate(wb_rows):
        key.cell(4+i,1,i+1); key.cell(4+i,2,rid)
        key.cell(4+i,3,cname); key.cell(4+i,4,cfg["source"])
    key.sheet_state = "hidden"
    book.save(workbook)

    with open(RESERVED_IDS, newline="", encoding="utf-8") as f:
        reg = [str(r["review_id"]) for r in csv.DictReader(f)]
    assert len(reg) == len(set(reg)), "reserved_ids.csv has duplicate review_ids"
    assert ids <= set(reg), "reserved_ids.csv is missing some drawn ids"
    assert_disjoint(ids, split, gold, stress, reserved_before, expected_n, "POST-WRITE")

    print(f"=== STAGE 2 COMPLETE ({lang}) ===")
    print(f"  workbook  : {workbook}  ({len(wb_rows)} rows, blind, shuffled)")
    print(f"  provenance: {ext_manifest}")
    print(f"  counts    : {ext_counts}")
    print(f"  reserved  : {RESERVED_IDS}  (now {len(reg)} ids total)")
    print("  assertions: PRE-WRITE + POST-WRITE disjoint(split/gold/stress/reserved) PASSED")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--lang", choices=["ja", "zh", "both"], default="both")
    ap.add_argument("--build", action="store_true",
                    help="STAGE 2: write workbook + manifests + reserved_ids "
                         "(requires the SIGNED pre-registration)")
    args = ap.parse_args()

    langs = ["zh", "ja"] if args.lang == "both" else [args.lang]
    split, gold, stress, reserved = load_exclusions()
    arm = load_arm_ids()
    full_excl = split | stress | reserved | arm

    for lang in langs:
        pool, frame_n = load_pool(lang, full_excl)
        hits, ungated, seen, _ = classify(lang, pool)
        if args.build:
            stage2(lang, hits, pool)
        else:
            excl_detail = [
                ("split_manifest (gold/pilot_draft/pilot_prompt)", len(split)),
                ("stress_candidate_manifest", len(stress)),
                ("reserved_ids.csv (en hardneg arm etc.)", len(reserved)),
                ("diagnostic-arm ids not already excluded", len(arm - (split | stress | reserved))),
                ("total distinct excluded", len(full_excl)),
            ]
            stage1(lang, hits, ungated, seen, pool, frame_n, excl_detail)

    if not args.build:
        print("\n(no text surfaced, no workbook, no LLM, nothing written — STAGE 1 only.)")


if __name__ == "__main__":
    main()
